# Shehadi Fino
# UWYO COSC 1010
# 11-
# Lab XX
# Lab Section: 
# Sources, people worked with, help given to: 
# your
# comments
# here

from openpyxl import Workbook
from openpyxl.styles import PatternFill

def mini_shuttle_pixel_art(file_name):
    wb = Workbook()
    ws = wb.active

    width_column = 4
    height_row = width_column * 6
    for col in range(1, 21):
        ws.column_dimensions[chr(64 + col)].width = width_column
    for row in range(1, 21):
        ws.row_dimensions[row].height = height_row

    pixel_art = {
        (2, 5): "FF0000", (2, 6): "FF0000", (2, 7): "FF0000", (2, 8): "FF0000", (2, 9): "FF0000",
        (3, 4): "FF0000", (3, 5): "FF0000", (3, 6): "0000FF", (3, 7): "0000FF", (3, 8): "FF0000", (3, 9): "FF0000", (3, 10): "FF0000",

        # Car body
        (4, 3): "FF0000", (4, 4): "FF0000", (4, 5): "FF0000", (4, 6): "0000FF", (4, 7): "0000FF", (4, 8): "FF0000", (4, 9): "FF0000", (4, 10): "FF0000", (4, 11): "FF0000",
        (5, 2): "FF0000", (5, 3): "FF0000", (5, 4): "FF0000", (5, 5): "FF0000", (5, 6): "0000FF", (5, 7): "0000FF", (5, 8): "FF0000", (5, 9): "FF0000", (5, 10): "FF0000", (5, 11): "FF0000", (5, 12): "FF0000",

        # Car tires
        (6, 3): "000000", (6, 4): "000000", (6, 5): "FF0000", (6, 6): "808080", (6, 7): "808080", (6, 8): "FF0000", (6, 9): "000000", (6, 10): "000000",
        (7, 3): "000000", (7, 4): "000000", (7, 5): "FFFFFF", (7, 6): "808080", (7, 7): "808080", (7, 8): "FFFFFF", (7, 9): "000000", (7, 10): "000000",
    
        }

    for (row, col), color in pixel_art.items():
        cell = ws.cell(row = row, column = col)
        cell.fill = PatternFill(start_color = color, end_color = color, fill_type = "solid")

    wb.save(file_name)
    print(f"Mini shuttle pixel art saved to {file_name}")  

mini_shuttle_pixel_art("mini_shuttle_pixel_art.xlsx")